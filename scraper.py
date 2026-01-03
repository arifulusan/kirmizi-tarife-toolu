#!/usr/bin/env python3
"""
Vodafone Tarife Scraper
Vodafone ve benzeri operatör sitelerinden tarife bilgilerini çekip Excel'e kaydeder.
"""

import asyncio
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from playwright.async_api import async_playwright


class TarifeScraper:
    """Web scraper for mobile tariff data."""
    
    def __init__(self, config_path: str = "config.json"):
        self.config = self._load_config(config_path)
        self.tariffs = []
        
    def _load_config(self, path: str) -> dict:
        """Load configuration from JSON file."""
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    async def scrape_vodafone(self, url: str) -> list[dict]:
        """Scrape tariff data from Vodafone website."""
        tariffs = []
        
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            print(f"🌐 Sayfa açılıyor: {url}")
            await page.goto(url, wait_until="networkidle")
            
            # Cookie popup'ı kapat
            try:
                reject_btn = page.locator("text=Reddet").first
                if await reject_btn.is_visible(timeout=3000):
                    await reject_btn.click()
                    await page.wait_for_timeout(500)
            except:
                pass
            
            # Sayfayı scroll yaparak tüm içeriği yükle
            print("📜 Sayfa scroll ediliyor...")
            for _ in range(8):
                await page.mouse.wheel(0, 1000)
                await page.wait_for_timeout(500)
            
            # Tarife verilerini çek
            print("📊 Tarife detayları çekiliyor (Bu işlem biraz zaman alabilir)...")
            
            # Önce temel konteynerları bulalım
            tariff_data = await page.evaluate("""
                async () => {
                    const results = [];
                    const containers = document.querySelectorAll('.css-1iqevk5');
                    
                    for (const container of containers) {
                        const headerEl = container.querySelector('p');
                        const categoryName = headerEl ? headerEl.textContent.trim() : 'Diğer Tarifeler';
                        
                        const selectBtns = Array.from(container.querySelectorAll('.chakra-button')).filter(b => b.textContent.includes('Tarifeyi seç'));
                        
                        for (const btn of selectBtns) {
                            const card = btn.closest('.css-1ir1t9b') || btn.closest('.css-0') || btn.parentElement.parentElement;
                            const text = card.innerText || '';
                            
                            // Temel bilgiler
                            const priceMatch = text.match(/(\\d{2,4})\\s*₺|₺\\s*(\\d{2,4})/);
                            const gbMatch = text.match(/(\\d+)\\s*GB/i);
                            const dkMatch = text.match(/(\\d+)\\s*DK/i);
                            const smsMatch = text.match(/(\\d+)\\s*SMS/i);
                            
                            if (priceMatch && gbMatch) {
                                const price = parseInt(priceMatch[1] || priceMatch[2]);
                                const gb = gbMatch[1];
                                const dk = dkMatch ? dkMatch[1] : '';
                                const sms = smsMatch ? smsMatch[1] : '';
                                
                                const lines = text.split('\\n').filter(l => l.trim());
                                let name = lines[0] || '';
                                if (name.length < 5 || /^\\d+$/.test(name.trim())) {
                                    for (const line of lines) {
                                        if (line.length > 5 && line.length < 50 && !line.includes('₺')) {
                                            name = line;
                                            break;
                                        }
                                    }
                                }

                                // Detayları gör butonunu bul ve tıkla
                                let noCommitmentPrice = '';
                                const detailBtn = Array.from(card.querySelectorAll('button')).find(b => b.textContent.includes('Detayları gör'));
                                
                                if (detailBtn) {
                                    detailBtn.click();
                                    // Modalın içeriğinin tamamen gelmesini bekle
                                    await new Promise(r => setTimeout(r, 1800));
                                    
                                    // Sayfadaki en son açılan veya görünür olan modalı yakala
                                    const modals = Array.from(document.querySelectorAll('[role="dialog"], .modal-content, [class*="Modal_content"]'));
                                    const modal = modals[modals.length - 1];
                                    
                                    if (modal) {
                                        const modalText = modal.innerText;
                                        // Kullanıcının belirttiği "Taahhütsüz Aylık Tarife Ücreti" keywordünü 
                                        // ve diğer varyasyonları (küçük/büyük harf, boşluklar) regex ile arıyoruz.
                                        const tcMatch = modalText.match(/Taahhütsüz.*?(?:ücreti|Ücreti)\s*:?\s*(\d{2,4})\s*TL/i) || 
                                                       modalText.match(/Taahhütsüz.*?(\d{2,4})\s*TL/i);
                                        
                                        if (tcMatch) {
                                            noCommitmentPrice = tcMatch[1];
                                        }
                                        
                                        // Kapatma butonu - Vodafone modal yapısına özel alternatifler
                                        const closeBtn = modal.querySelector('button[aria-label="Close"]') || 
                                                       Array.from(modal.querySelectorAll('button, span, i')).find(b => 
                                                            b.innerText === '✕' || b.innerText === 'X' || 
                                                            b.innerText.includes('Kapat') || 
                                                            b.className.includes('close')
                                                       );
                                        if (closeBtn) closeBtn.click();
                                        await new Promise(r => setTimeout(r, 800));
                                    }
                                }
                                
                                results.push({
                                    category: categoryName,
                                    name: name.trim().substring(0, 60),
                                    gb: gb,
                                    minutes: dk,
                                    sms: sms,
                                    price: price,
                                    no_commitment_price: noCommitmentPrice,
                                    provider: 'Vodafone'
                                });
                            }
                        }
                    }
                    return results;
                }
            """)
            
            # Fiyata göre sıralama (Python tarafında yapalım daha temiz olur)
            from collections import defaultdict
            grouped = defaultdict(list)
            for t in tariff_data:
                grouped[t['category']].append(t)
            
            tariffs = []
            for category in grouped:
                grouped[category].sort(key=lambda x: x['price'])
                tariffs.extend(grouped[category])
            
            await browser.close()
            
        print(f"✅ {len(tariffs)} tarife bulundu")
        return tariffs

    async def scrape_turkcell(self, url: str) -> list[dict]:
        """Scrape tariff data from Turkcell website."""
        tariffs = []
        
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            print(f"🌐 Sayfa açılıyor: {url}")
            await page.goto(url, wait_until="networkidle")
            
            # Popupları kapat
            try:
                # Cookie kabul
                accept_btn = page.locator("text=Kabul Et").first
                if await accept_btn.is_visible(timeout=5000):
                    await accept_btn.click()
                
                # Bildirim uyarısı (Daha Sonra)
                later_btn = page.locator("#btn-later").first
                if await later_btn.is_visible(timeout=3000):
                    await later_btn.click()
            except:
                pass
            
            # Sayfayı scroll yaparak tüm içeriği yükle
            print("📜 Sayfa scroll ediliyor...")
            for _ in range(10):
                await page.mouse.wheel(0, 1000)
                await page.wait_for_timeout(500)
            
            # Tarife verilerini çek
            print("📊 Turkcell tarifeleri çekiliyor...")
            
            tariff_data = await page.evaluate("""
                async () => {
                    const results = [];
                    // Turkcell kart seçici
                    const cards = document.querySelectorAll('.molecules-teasy-card_m-teasy-card__Ly4fG');
                    
                    for (const card of cards) {
                        try {
                            const titleEl = card.querySelector('.molecules-teasy-card_m-teasy-card__title__h0CO1');
                            const name = titleEl?.textContent?.trim() || 'Turkcell Tarife';
                            const badgeEl = card.querySelector('.molecules-teasy-card_m-teasy-card__badge__nd1eJ');
                            const badgeText = badgeEl?.textContent?.trim() || '';
                            
                            // Kategori belirleme mantığı
                            let category = 'Diğer Tarifeler';
                            const lowerName = name.toLowerCase();
                            const lowerBadge = badgeText.toLowerCase();
                            
                            if (lowerBadge.includes('online')) {
                                category = "Online'a Özel Tarifeler";
                            } else if (lowerBadge.includes('platinum') || lowerName.includes('platinum')) {
                                category = "Platinum Tarifeleri";
                            } else if (lowerBadge.includes('gnç') || lowerName.includes('gnç')) {
                                category = "GNÇ Tarifeleri";
                            } else if (badgeText) {
                                category = badgeText + " Tarifeleri";
                            }
                            
                            const gbText = card.querySelector('.molecules-teasy-card_m-teasy-card__text__container__UY7Ei')?.textContent?.trim() || '';
                            const dkText = card.querySelector('.molecules-teasy-card_m-teasy-card__subtext__3SrTQ')?.textContent?.trim() || '';
                            const priceText = card.querySelector('.atom-price_a-price__7lMAa span:first-child')?.textContent?.trim() || '';
                            
                            // Sayılar temizle
                            const gb = gbText.match(/(\\d+)/)?.[1] || '';
                            const price = parseInt(priceText.replace(/\\D/g, '')) || 0;
                            const dk = dkText.match(/(\\d+)/)?.[1] || '';
                            
                            let sms = '';
                            
                            // Detay modalını açıp SMS bilgisi almayı dene
                            const detailBtn = Array.from(card.querySelectorAll('button, a')).find(el => el.textContent.includes('DETAY'));
                            if (detailBtn) {
                                detailBtn.click();
                                await new Promise(r => setTimeout(r, 1200));
                                
                                const modal = document.querySelector('.ant-modal-content');
                                if (modal) {
                                    const modalText = modal.innerText;
                                    const smsMatch = modalText.match(/(\\d+)\\s*SMS/i);
                                    if (smsMatch) sms = smsMatch[1];
                                    
                                    // Modalı kapat
                                    const closeBtn = Array.from(modal.querySelectorAll('button, span, div')).find(el => el.textContent.trim() === 'Vazgeç' || el.classList.contains('ant-modal-close'));
                                    if (closeBtn) closeBtn.click();
                                    await new Promise(r => setTimeout(r, 500));
                                }
                            }
                            
                            results.push({
                                category: category,
                                name: name,
                                gb: gb,
                                minutes: dk,
                                sms: sms,
                                price: price,
                                no_commitment_price: '',
                                provider: 'Turkcell'
                            });
                        } catch (e) {
                            console.error('Card extraction error:', e);
                        }
                    }
                    return results;
                }
            """)
            
            tariffs = sorted(tariff_data, key=lambda x: x['price'])
            await browser.close()
            
        print(f"✅ {len(tariffs)} Turkcell tarifesi bulundu")
        return tariffs

    async def scrape_turkcell_mevcut(self, url: str) -> list[dict]:
        """Scrape Turkcell existing customer tariffs."""
        tariffs = []
        
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36")
            page = await context.new_page()
            
            print(f"🌐 Liste sayfası açılıyor: {url}")
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            
            # Kartların yüklenmesini bekle
            try:
                await page.wait_for_selector('a.molecule-dynamic-card_linkDecoration__cDpXS', timeout=20000)
            except:
                print("⚠️ Uyarı: Kartlar beklenen sürede yüklenmedi, yine de devam ediliyor.")

            # Popupları kapatmayı dene
            try:
                accept_btn = page.locator("text=Kabul Et").first
                if await accept_btn.is_visible(timeout=3000):
                    await accept_btn.click()
            except: pass
            
            # Sayfayı scroll yaparak tüm içeriği yükle
            for _ in range(3):
                await page.mouse.wheel(0, 1500)
                await page.wait_for_timeout(800)
            
            # Linkleri topla
            tariff_links = await page.evaluate("""
                () => {
                    const links = Array.from(document.querySelectorAll('a.molecule-dynamic-card_linkDecoration__cDpXS'))
                                       .map(a => a.href);
                    return [...new Set(links)]; // Tekrar edenleri temizle
                }
            """)
            
            if not tariff_links:
                print("❌ Hata: Hiç tarife linki bulunamadı. Seçici değişmiş olabilir.")
                await browser.close()
                return []

            print(f"🔗 {len(tariff_links)} adet tarife linki bulundu. Detaylar çekiliyor...")
            
            detail_page = await context.new_page()
            
            for i, link in enumerate(tariff_links, 1):
                try:
                    print(f"📝 ({i}/{len(tariff_links)}) taranıyor: {link}")
                    await detail_page.goto(link, wait_until="domcontentloaded", timeout=30000)
                    # Çok hızlı gidince bloklanmamak için kısa bekleme
                    await detail_page.wait_for_timeout(1000) 
                    
                    data = await detail_page.evaluate("""
                        () => {
                            const name = document.querySelector('h1')?.textContent?.trim() || 
                                         document.querySelector('h2')?.textContent?.trim() || 'Turkcell Tarife';
                            
                            let gb = '', dk = '', sms = '';
                            // Daha geniş bir seçici grubu
                            const elements = Array.from(document.querySelectorAll('h1, h2, h3, p, div[class*="packageName"]'));
                            elements.forEach(el => {
                                const txt = el.innerText.toUpperCase();
                                if (/^\\d+\\s*GB$/i.test(txt) || (txt.includes('GB') && txt.length < 15)) {
                                    gb = txt.replace('GB', '').trim();
                                } else if (txt.includes('DK') && txt.length < 15) {
                                    dk = txt.replace('DK', '').trim();
                                } else if (txt.includes('SMS') && txt.length < 15) {
                                    sms = txt.replace('SMS', '').trim();
                                }
                            });
                            
                            let price = 0;
                            let noCommitmentPrice = 0;
                            
                            // Fiyatları sayfa metni içinde ara
                            const bodyText = document.body.innerText;
                            
                            // Yıllık Taahhütlü Fiyat
                            const annualMatch = bodyText.match(/Yıllık\\s*Abonelik.*?(\\d+)\\s*TL/is);
                            if (annualMatch) price = parseInt(annualMatch[1]);
                            
                            // Aylık Taahhütsüz Fiyat
                            const monthlyMatch = bodyText.match(/Aylık\\s*Abonelik.*?(\\d+)\\s*TL/is);
                            if (monthlyMatch) noCommitmentPrice = parseInt(monthlyMatch[1]);
                            
                            // Alternatif: Radyo butonlarından çekmeyi dene (görseldeki yapı)
                            const priceLabels = Array.from(document.querySelectorAll('label, .ant-radio-wrapper'));
                            priceLabels.forEach(label => {
                                const lText = label.innerText.toUpperCase();
                                const pMatch = label.innerText.match(/(\\d+)\\s*TL/i);
                                if (pMatch) {
                                    const val = parseInt(pMatch[1]);
                                    if (lText.includes('YILLIK')) price = val;
                                    else if (lText.includes('AYLIK')) noCommitmentPrice = val;
                                }
                            });

                            return {
                                name: name,
                                gb: gb,
                                minutes: dk,
                                sms: sms,
                                price: price,
                                no_commitment_price: noCommitmentPrice
                            };
                        }
                    """)
                    
                    if data['price'] == 0 and data['no_commitment_price'] > 0:
                        data['price'] = data['no_commitment_price'] # Fallback

                    category = 'Diğer Tarifeler'
                    lowerName = data['name'].toLowerCase()
                    if 'platinum' in lowerName: category = 'Platinum Tarifeleri'
                    elif 'star' in lowerName: category = 'Star Tarifeleri'
                    elif 'esneyen' in lowerName: category = 'Esneyen Tarifeler'
                    elif 'gnç' in lowerName: category = 'GNÇ Tarifeleri'
                    
                    tariffs.append({
                        'category': category,
                        'name': data['name'],
                        'gb': data['gb'],
                        'minutes': data['minutes'],
                        'sms': data['sms'],
                        'price': data['price'],
                        'no_commitment_price': data['no_commitment_price'],
                        'provider': 'Turkcell (Mevcut)'
                    })
                    
                except Exception as e:
                    print(f"⚠️ Hata (Atlanıyor - {link}): {str(e)}")
                    continue
            
            await browser.close()
            
        # Fiyata göre sırala
        tariffs = sorted(tariffs, key=lambda x: x['price'] if x['price'] > 0 else 9999)
        print(f"✅ Bitti: {len(tariffs)} Turkcell Mevcut tarifesi çekildi.")
        return tariffs
    
    def save_to_excel(self, tariffs: list[dict], output_path: str):
        """Save tariff data to Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Tarifeler"
        
        # Başlık stili
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="E60000", end_color="E60000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlıklar
        headers = ["Kategori", "Paket Adı", "İnternet (GB)", "Dakika", "SMS", "Fiyat (₺/ay)", "Taahhütsüz Fiyat (₺/ay)", "Kaynak", "Tarih"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Veri satırları
        today = datetime.now().strftime("%Y-%m-%d %H:%M")
        for row, tariff in enumerate(tariffs, 2):
            provider = tariff.get('provider', 'Vodafone')
            ws.cell(row=row, column=1, value=tariff.get('category', '')).border = thin_border
            ws.cell(row=row, column=2, value=tariff.get('name', '')).border = thin_border
            ws.cell(row=row, column=3, value=tariff.get('gb', '')).border = thin_border
            ws.cell(row=row, column=4, value=tariff.get('minutes', '')).border = thin_border
            ws.cell(row=row, column=5, value=tariff.get('sms', '')).border = thin_border
            ws.cell(row=row, column=6, value=tariff.get('price', '')).border = thin_border
            ws.cell(row=row, column=7, value=tariff.get('no_commitment_price', '')).border = thin_border
            ws.cell(row=row, column=8, value=provider).border = thin_border
            ws.cell(row=row, column=9, value=today).border = thin_border
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 18
        
        wb.save(output_path)
        print(f"💾 Excel dosyası kaydedildi: {output_path}")
    
    async def run(self):
        """Run the scraper for all configured URLs."""
        all_tariffs = []
        
        for site in self.config.get('urls', []):
            name = site.get('name', 'Unknown')
            url = site.get('url', '')
            
            print(f"\n{'='*50}")
            print(f"📱 {name} tarifelerini çekiyor...")
            print(f"{'='*50}")
            
            if 'vodafone' in url.lower():
                tariffs = await self.scrape_vodafone(url)
                all_tariffs.extend(tariffs)
            else:
                print(f"⚠️  {name} için scraper henüz eklenmedi")
        
        if all_tariffs:
            output_path = self.config.get('output_file', 'tarifeler.xlsx')
            self.save_to_excel(all_tariffs, output_path)
            print(f"\n🎉 Toplam {len(all_tariffs)} tarife çekildi ve kaydedildi!")
        else:
            print("\n❌ Hiç tarife bulunamadı!")


async def main():
    scraper = TarifeScraper()
    await scraper.run()


if __name__ == "__main__":
    asyncio.run(main())
